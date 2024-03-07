__author__ = ["David Cardona-Vasquez"]
__copyright__ = "Copyright 2022, Graz University of Technology"
__credits__ = ["David Cardona-Vasquez"]
__license__ = "MIT"
__maintainer__ = "David Cardona-Vasquez"
__status__ = "Development"

import glob
import numpy as np
import os

import openpyxl
import pyomo.environ as pyo
import pandas as pd
import yaml


def create_complete_model(input_data:dict):
    """
    This function creates the complete optimization model using the provided input data. Returns a Pyomo concrete model
    :param input_data:
    :return:
    """
    model = pyo.ConcreteModel(name="(Com)")

    the_gen = input_data['thermal_generators']
    wnd_gen = input_data['wind_generators']
    p = input_data['periods']
    pVC_g = input_data['vc_gen']
    pVC_nsp = input_data['vc_nsp']
    pCapFac_w = input_data['cf_wind']
    pMinProd_g = input_data['minprod']
    pMaxProd_g = input_data['maxprod']
    pDemand_p = input_data['demand']


    model.g = input_data['generators']
    model.t = pyo.Set(within=model.g, initialize=the_gen)
    model.w = pyo.Set(within=model.g, initialize=wnd_gen)
    model.p = pyo.RangeSet(p)

    model.pVC_g = pyo.Param(model.g, initialize=pVC_g)
    model.pVC_nsp = pyo.Param(initialize=pVC_nsp)
    model.pCapFac_w = pyo.Param(model.p, model.w, initialize=pCapFac_w, domain=pyo.NonNegativeReals)
    model.pMinProd_g = pyo.Param(model.g, initialize=pMinProd_g)
    model.pMaxProd_g = pyo.Param(model.g, initialize=pMaxProd_g)
    model.pDemand_p = pyo.Param(model.p, initialize=pDemand_p, domain=pyo.NonNegativeReals)

    model.vGen = pyo.Var(model.g, model.p, domain=pyo.NonNegativeReals)
    model.vNSP = pyo.Var(model.p, domain=pyo.NonNegativeReals)


    def eNSP_rule(mdl, i):
        return mdl.pDemand_p[i] >= mdl.vNSP[i]
    model.eNSP = pyo.Constraint(model.p, rule=eNSP_rule)

    def eBalance_rule(mdl, i):
        return mdl.pDemand_p[i] == sum(mdl.vGen[aux_g, i] for aux_g in mdl.g) + \
                mdl.vNSP[i]
    model.eBalance = pyo.Constraint(model.p, rule=eBalance_rule)

    def eMinProd_rule(mdl, g, i):
        return mdl.pMinProd_g[g] <= mdl.vGen[g, i]
    model.eMinProd = pyo.Constraint(model.g, model.p, rule=eMinProd_rule)

    def eMaxProd_rule(mdl, g, i):
        if g in mdl.w:
            return mdl.pMaxProd_g[g]*mdl.pCapFac_w[i, g] >= mdl.vGen[g, i]
        else:
            return mdl.pMaxProd_g[g] >= mdl.vGen[g, i]
    model.eMaxProd = pyo.Constraint(model.g, model.p, rule=eMaxProd_rule)

    def eCost_rule(mdl, i):
        return sum(mdl.vGen[g, i]*mdl.pVC_g[g] for g in mdl.g) + mdl.vNSP[i]*mdl.pVC_nsp
    model.vCost = pyo.Expression(model.p, rule=eCost_rule)

    def eObjective_rule(mdl):
        return sum(sum(mdl.vGen[g, i]*mdl.pVC_g[g] for g in mdl.g) + mdl.vNSP[i]*mdl.pVC_nsp for i in model.p)
    model.z = pyo.Objective(rule=eObjective_rule, sense=pyo.minimize)

    model.rc = pyo.Suffix(direction=pyo.Suffix.IMPORT)
    model.dual = pyo.Suffix(direction=pyo.Suffix.IMPORT)


    return model



def basis_summarization(sPath:str, basis:list) -> pd.DataFrame:
    """
    Receives a path to look for the basis and a list of the basis to consider;
    returns the input data (demand and capacity factor) for each basis
    :param sPath: folder with the Excel files with the basis are
    :param basis: string list with the basis to consider for the summarization
    :return: DataFrame with the demand and capacity factor of each basis
    """

    dfs = []
    for k in basis:
        auxPath = os.path.join(sPath, str(k)+".xlsx")
        wsDemand = pd.read_excel(auxPath, "demand", index_col=0)
        wsCapFactors = pd.read_excel(auxPath, "cap_factors", index_col=0)

        df_aux = pd.merge(left=wsDemand, right=wsCapFactors, how='inner', left_index=True, right_index=True)
        df_aux.drop(columns=['generator'], inplace=True)
        df_aux['basis'] = str(k)

        dfs.append(df_aux.copy())

    dfAgg = pd.concat(dfs)
    dfAgg.reset_index(drop=True, inplace=True)

    return dfAgg


def data_load(sPath:str):
    """
    This function is used to load the data from the Excel files which have the parametrization of each 'basis' run
    of the power system
    :param file:
    :return:
    """

    wsConfig = pd.read_excel(sPath,"config", header=None, index_col=0)
    wsThermal = pd.read_excel(sPath, "thermal", index_col=0)
    wsVRES = pd.read_excel(sPath, "vres", index_col=0)
    wsDemand = pd.read_excel(sPath, "demand", index_col=0)
    wsCapFactors = pd.read_excel(sPath, "cap_factors", index_col=0)

    input_data = dict()
    input_data['periods'] = wsConfig.loc['periods'].values[0]
    input_data['vc_nsp'] = wsConfig.loc['vc_nsp'].values[0]

    # get the data from the thermal generators
    the_gen = list(wsThermal.index)
    input_data['thermal_generators'] = the_gen

    # get the data from the wind generators
    wnd_gen = list(wsVRES.index)
    input_data['wind_generators'] = wnd_gen

    the_gen.extend(wnd_gen)
    input_data['generators'] = the_gen

    # use the data from thermal generators for its parameters
    input_data['rmpup'] = dict()
    input_data['rmpdn'] = dict()
    input_data['maxprod'] = dict()
    input_data['minprod'] = dict()
    input_data['vc_gen'] = dict()
    for idx, r in wsThermal.iterrows():
        input_data['rmpup'][idx] = r['rmp_up']
        input_data['rmpdn'][idx] = r['rmp_dn']

        input_data['minprod'][idx] = r['min_cap']
        input_data['maxprod'][idx] = r['max_cap']
        input_data['vc_gen'][idx] = r['vc']

    # use the data from wind generators for its parameters
    for idx, r in wsVRES.iterrows():
        input_data['minprod'][idx] = r['min_cap']
        input_data['maxprod'][idx] = r['max_cap']
        input_data['vc_gen'][idx] = r['vc']

    # get data for capacity factors and demand
    input_data['demand'] = dict()
    input_data['cf_wind'] = dict()
    for idx, r in wsDemand.iterrows():
        input_data['demand'][idx] = r['demand']

    for idx, r in wsCapFactors.iterrows():
        input_data['cf_wind'][(idx, r['generator'])] = r['cap_factor']


    return input_data


def basis_execution(reference='data/complete_single_node.xlsx', folder='.', file='config.xlsx', solver='gurobi'):
    """
    This function runs an optimization model considering the parametrized 'basis' in the yaml file. Each 'basis'must
    have an associated Excel file with the power system configuration to be used.
    :param reference:
    :param folder:
    :param file:
    :param solver:
    :return:
    """

    df_config = pd.read_excel(os.path.join(folder, file))
    wb_model = openpyxl.load_workbook(filename=reference)

    if 'cap_factors' in wb_model.sheetnames:
        wb_model.remove(wb_model['cap_factors'])
    wb_model.create_sheet('cap_factors')
    ws_cf = wb_model['cap_factors']
    ws_cf.cell(row=1, column=1).value = 'period'
    ws_cf.cell(row=1, column=2).value = 'generator'
    ws_cf.cell(row=1, column=3).value = 'cap_factor'
    ws_cf.cell(row=2, column=1).value = 1
    ws_cf.cell(row=2, column=2).value = 'w1'


    if 'demand' in wb_model.sheetnames:
        wb_model.remove(wb_model['demand'])
    wb_model.create_sheet('demand')
    ws_dem = wb_model['demand']
    ws_dem.cell(row=1, column=1).value = 'period'
    ws_dem.cell(row=1, column=2).value = 'demand'
    ws_dem.cell(row=2, column=1).value = 1

    ws_config = wb_model['config']
    ws_config.cell(row=1, column=2).value = 1

    if not os.path.exists(folder):
        os.mkdir(folder)

    models = dict()
    for idx, r in df_config.iterrows():
        cent_dem = r['centroid_demand']
        cent_cf = r['centroid_cf']
        name = r['basis']
        weight = r['weight']

        ws_cf.cell(row=2, column=3).value = cent_cf
        ws_dem.cell(row=2, column=2).value = cent_dem

        # this code is redundant but is to make use of already existing code
        # to load the data for the model run from an Excel file
        sPath = os.path.join(folder, name+'.xlsx')
        wb_model.save(sPath)
        data = data_load(sPath)
        model = create_complete_model(data)
        solved = pyo.SolverFactory(solver, solver_io="lp")
        res = solved.solve(model)
        aux_res = dict()

        aux_res['z'] = pyo.value(model.z) * int(weight)
        aux_res['model'] = model
        aux_res['weight'] = int(weight)

        models[name] = aux_res

    return models


def run_complete_case(case='complete_1.xlsx', folder='data', solver='gurobi'):
    """
    This is used to run a complete model (without any aggregation) using the parametrization provided in the Excel file.
    The main differente with the 'basis' run is that it corresponds to only one complete run and thus it does not
    require any kind of weighting of the results.
    :return:
    """

    sPath = os.path.join(folder, case)
    data = data_load(sPath)
    model = create_complete_model(data)
    solved = pyo.SolverFactory(solver, solver_io="lp")
    res = solved.solve(model)

    return model


def idx_match(row:pd.Series, df:pd.DataFrame):
    idxs = []
    for idx, r in df.iterrows():
        if all(row == r):
            idxs.append(idx)

    return idxs

def get_centroids(b_idx:dict, data:pd.DataFrame, cols=[0, 1], col_names=['demand', 'cap_factor']):

    b_cent = []
    names = []
    i=1
    for b in b_idx.values():
        aux = data.iloc[b, cols].apply(np.average, axis=0).values
        b_cent.append(aux)
        names.append('bs' + str(i))
        i = i + 1

    df_cent = pd.DataFrame(b_cent)
    df_cent.columns = col_names
    df_cent['basis'] = names

    return df_cent



def export_complete_solution(mdl: pyo.ConcreteModel, folder: str, hourly=True):

    sln = []

    sln_dict = dict()
    l_sln_hourly = []
    if not os.path.exists(folder):
        os.mkdir(folder)

    sln_dict['of_value'] = pyo.value(mdl.z)
    sln_dict['thermal'] = 0
    sln_dict['renewable'] = 0
    sln_dict['nsp'] = 0

    # this code is only for illustrative purposes as it is not
    # the most efficient way to obtain data from Pyomo
    for v in mdl.component_objects(ctype=pyo.Var):
        aux_var = {'var': str(v)}
        aux_d = {}
        for idx in v:
            aux_d[idx] = pyo.value(v[idx])
            if str(v) == 'vGen':
                g, p = idx
                if g == 'w1':
                    sln_dict['renewable'] = sln_dict['renewable'] + aux_d[idx]
                else:
                    sln_dict['thermal'] = sln_dict['thermal'] + aux_d[idx]
            elif str(v) == 'vNSP':
                sln_dict['nsp'] = sln_dict['nsp'] + aux_d[idx]
            else:
                raise Exception("Unknown variable in model!")

        aux_var['values'] = aux_d

        sln.append(aux_var)

    for elem in sln:
        df = pd.DataFrame.from_dict(elem['values'], orient='index', columns=[str(elem['var'])])

        if not df.index.is_numeric():
            df.index = pd.MultiIndex.from_tuples(df.index)

        df.reset_index(drop=False, inplace=True)
        df.to_excel(os.path.join(folder, str(elem['var'])+'.xlsx'), index=False)
        if 'level_1' in df.columns:
            df = df.pivot(index='level_1', columns='level_0', values='vGen').reset_index()
            df.rename(columns={'level_1': 'period'}, inplace=True)
        else:
            df.rename(columns={'index': 'period'}, inplace=True)
        l_sln_hourly.append(df)

    df_hourly = None
    if len(l_sln_hourly) > 0:
        df_hourly = l_sln_hourly[0]
        for i in range(1, len(l_sln_hourly)):
            df_hourly = df_hourly.merge(right=l_sln_hourly[i], on='period')

    df_complete = pd.DataFrame.from_dict(sln_dict, orient='index').rename(columns={0: 'complete'})
    df_complete.to_excel(os.path.join(folder, 'results.xlsx'))

    return df_complete, df_hourly


def export_aggregated_solution(mdls: dict, folder: str):

    # this code is only for illustrative purposes as it is not
    # the most efficient way to obtain data from Pyomo

    if not os.path.exists(folder):
        os.mkdir(folder)

    l_names = []
    l_of = []
    l_thermal = []
    l_renewable = []
    l_nsp = []
    l_weight = []
    for k, v in mdls.items():
        mdl = v['model']
        l_names.append(k)
        l_of.append(v['z'])
        l_renewable.append(pyo.value(mdl.vGen[('w1', 1)])*v['weight'])
        l_thermal.append(pyo.value(mdl.vGen[('t1', 1)])*v['weight'])
        l_nsp.append(pyo.value(mdl.vNSP[1])*v['weight'])
        l_weight.append(v['weight'])

    df_results = pd.DataFrame({'basis': l_names,
                               'of_value': l_of,
                               'thermal': l_thermal,
                               'renewable': l_renewable,
                               'nsp': l_nsp,
                               'weight': l_weight})
    df_results.to_excel(os.path.join(folder, 'results.xlsx'), index=False)

    return df_results


def export_model_comparison(df_complete: pd.DataFrame, df_aggregated: pd.DataFrame) -> pd.DataFrame:
    """
    Receives a complete model and an aggregated one that intends to approximate it.
    :param aggregated: list of models returned by export_aggregated_solution
    :param complete: complete model returned by run_complete_case
    :return: dataframe with results comparison
    """

    decision_vars = ['decision_vars', 8736*3, df_aggregated.shape[0]*3]

    agg_sum = df_aggregated.loc[:, ['of_value', 'thermal', 'renewable', 'nsp']].sum()
    df_complete = df_complete.copy()
    df_complete.insert(1, 'aggregated', agg_sum)
    df_complete.reset_index(drop=False, inplace=True)
    df_complete.rename(columns={'index': 'result'}, inplace=True)

    df_complete.loc[len(df_complete)] = decision_vars

    df_complete['delta'] = df_complete['complete'] - df_complete['aggregated']
    df_complete['rel_delta'] = 1 - df_complete['aggregated']/df_complete['complete']


    return df_complete


def extract_duals(model, folder: str):

    aux_d = {}
    for k, v in model.dual.items():
        aux_name = k.parent_component().local_name
        if isinstance(k.index(), tuple):
            aux_name = aux_name + "_" + pyo.value(k.index()[0])

        try:
            aux_d[aux_name]
        except KeyError as e:
            aux_d[aux_name] = {}
        if isinstance(k.index(), tuple):
            aux_d[aux_name][pyo.value(k.index()[1])] = v
        else:
            aux_d[aux_name][pyo.value(k.index())] = v

    df_duals = pd.DataFrame.from_dict(aux_d['eBalance'], orient='index').reset_index(drop=False)
    df_duals.rename(columns={0: 'eBalance'}, inplace=True)

    for k, v in aux_d.items():
        if not k == 'eBalance':
            df_aux = pd.DataFrame.from_dict(aux_d[k], orient='index').reset_index(drop=False).rename(columns={0: k})
            df_duals = pd.merge(left=df_duals, right=df_aux, left_on='index', right_on='index', how='left')

    df_duals.rename(columns={'index': 'period'}, inplace=True)
    df_duals.fillna(0, inplace=True)

    if not os.path.exists(folder):
        os.mkdir(folder)

    df_duals.to_excel(os.path.join(folder, folder+"_duals.xlsx"), index=False)

    return df_duals



def generate_config(df_centroids: pd.DataFrame, folder: str, data_folder=True):
    """
    Code to export the centroids in df_centroids to the config_auto.xlsx, you can do it by hand to tweak it to your own clustering
    :param df_centroids:
    :param folder:
    :return:
    """

    if data_folder:
        prev_files = glob.glob(os.path.join('data',folder,'bs_auto*.xlsx'))
    else:
        prev_files = glob.glob(os.path.join(folder, 'bs_auto*.xlsx'))

    for f in prev_files:
        os.remove(f)

    # copying the dataframe
    df_config_auto = df_centroids.copy()
    df_config_auto['basis'] = ['bs_auto' + str(i) for i in range(1, df_config_auto.shape[0] + 1)]

    # renaming the columns
    df_config_auto.rename(columns={'cap_factor': 'centroid_cf', 'demand': 'centroid_demand'}, inplace=True)
    # changing the order of the centroids
    df_config_auto = df_config_auto.iloc[:, [3, 1, 0, 2]]
    # exporting the file

    if data_folder:
        df_config_auto.to_excel(os.path.join('data',folder, 'config_auto.xlsx'), index=False)
    else:
        df_config_auto.to_excel(os.path.join(folder, 'config_auto.xlsx'), index=False)

    return df_config_auto




if __name__ == '__main__':
    results_complete = run_complete_case(folder='data', case='complete_single_node.xlsx')
    # df_complete, df_hourly = export_complete_solution(results_complete, '.')
    df_duals = extract_duals(results_complete, 'complete_single_node')
    # models = basis_execution(folder='data/aggregated_single')
    # df_agg = export_aggregated_solution(models, 'aggregated_single')
    # df_comparison = export_model_comparison(df_complete, df_agg)
